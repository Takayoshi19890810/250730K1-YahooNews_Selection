from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import Workbook
import time
import os

# Google Sheets & Drive Libraries
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive

# --- 設定 ---
# 読み込むGoogle SheetsのURL
SPREADSHEET_URL = "https://docs.google.com/spreadsheets/d/1nphpu1q2cZuxJe-vYuOliw1azxqKKlzt6FFGNEJ76sw/edit?gid=0#gid=0"
# Google DriveにアップロードするフォルダのID
# https://drive.google.com/drive/folders/1MjNzGR57vsLtjbBJAZl06BKqZALYjGUO?usp=sharing からIDを抽出
# URLの 'folders/' の後に続く部分がIDです
DRIVE_FOLDER_ID = "1MjNzGR57vsLtjbBJAZl06BKqZALYjGUO"
# サービスアカウントキーのJSONファイル名（GitHub ActionsではSecretsから動的に生成）
SERVICE_ACCOUNT_FILE = 'service_account.json'

# ファイル設定
TODAY_STR = datetime.now().strftime("%y%m%d")
OUTPUT_FILE = f"{TODAY_STR}.xlsx" # ローカルで一時的に作成するExcelファイル名

def parse_relative_time(text):
    """相対的な時間表現（例: '5分前'）をdatetimeオブジェクトに変換する"""
    now = datetime.now()
    try:
        if "分前" in text:
            return now - timedelta(minutes=int(text.replace("分前", "").strip()))
        elif "時間前" in text:
            return now - timedelta(hours=int(text.replace("時間前", "").strip()))
        elif "日前" in text:
            return now - timedelta(days=int(text.replace("日前", "").strip()))
        elif "秒前" in text:
            return now - timedelta(seconds=int(text.replace("秒前", "").strip())) # 秒前も考慮
    except ValueError:
        pass # 変換できない場合は後続のtry-exceptで処理
    # 絶対的な日付形式（例: '2023年7月29日'）も考慮
    try:
        # よくある日付形式のパターン
        for fmt in ("%Y年%m月%d日 %H時%M分", "%Y/%m/%d %H:%M", "%y/%m/%d %H:%M", "%Y年%m月%d日", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"): # ★追加
            if "時" not in text and "分" not in text and "日" in fmt: # 時間情報がない場合の考慮
                return datetime.strptime(text.split(' ')[0], fmt).replace(hour=0, minute=0, second=0)
            else:
                return datetime.strptime(text, fmt)
    except ValueError:
        pass
    return now # どの形式にもマッチしない場合は現在時刻を返す

def format_datetime(dt):
    """datetimeオブジェクトを指定のフォーマット文字列に変換する"""
    return dt.strftime("%y/%m/%d %H:%M")

def get_news_pages(base_url, driver):
    """ニュース記事の本文を複数ページにわたって取得する"""
    page_contents = []
    page = 1
    last_content = ""

    while True:
        url = base_url if page == 1 else f"{base_url}?page={page}"
        driver.get(url)
        time.sleep(2) # ページの読み込みを待つ

        try:
            # 記事本文のある要素を特定 (Yahoo!ニュースのHTML構造変更に注意)
            article_element = driver.find_element(By.TAG_NAME, "article")
            paragraphs = article_element.find_elements(By.TAG_NAME, "p")
            content = "\n".join([p.text for p in paragraphs if p.text.strip()]).strip()

            if not content or content == last_content:
                break # 内容が空または前ページと同じ場合は終了

            page_contents.append(content)
            last_content = content
            page += 1
        except Exception as e:
            # print(f"記事ページ取得エラー ({url}): {e}")
            break # 記事要素が見つからない場合は終了

    # タイトルを取得するために再度最初のURLにアクセス（必要であれば）
    driver.get(base_url)
    time.sleep(1)
    try:
        title = driver.title.replace(" - Yahoo!ニュース", "")
    except:
        title = "タイトル取得失敗"
    return title, base_url, page_contents

def get_comments_pages(base_url, driver):
    """ニュース記事のコメントを複数ページにわたって取得する"""
    comments_data = []
    page = 1
    last_comments_joined = ""
    # Yahoo!ニュースのコメントページのURL構造
    article_id = base_url.rstrip("/").split("/")[-1]
    base_comment_url = f"https://news.yahoo.co.jp/articles/{article_id}/comments"

    while True:
        comment_url = base_comment_url if page == 1 else f"{base_comment_url}?page={page}"
        driver.get(comment_url)
        time.sleep(2) # コメントページの読み込みを待つ

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        # コメント要素のCSSセレクタ (Yahoo!ニュースのHTML構造変更に注意)
        comment_elements = soup.find_all('article', class_='sc-169yn8p-3') # コメントの親要素のクラス名

        page_comments = []
        for comment_article in comment_elements:
            comment_p = comment_article.find('p', class_='sc-169yn8p-10') # コメント本文のクラス名
            comment_text = comment_p.text.strip() if comment_p else ''
            user_a = comment_article.find('a', class_='sc-169yn8p-7') # ユーザー名のクラス名
            user_name = user_a.text.strip() if user_a else ''
            time_a = comment_article.find('a', class_='sc-169yn8p-9') # 投稿日時のクラス名
            raw_time = time_a.text.strip() if time_a else ''

            dt = parse_relative_time(raw_time)
            formatted_time = format_datetime(dt)
            page_comments.append((comment_text, formatted_time, user_name))

        joined_current_page = "\n".join([c[0] for c in page_comments])
        if not page_comments or joined_current_page == last_comments_joined:
            break # コメントがない、または前ページと同じ内容の場合は終了

        last_comments_joined = joined_current_page
        comments_data.extend(page_comments)
        page += 1

    return comments_data if comments_data else [("コメントなし", "", "")]

def get_urls_from_google_sheet(spreadsheet_url):
    """Google SheetsからURLと投稿日時を読み込む"""
    try:
        # サービスアカウント認証の設定
        # Google Drive APIスコープも追加（PyDrive2で利用するため）
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, scope)
        client = gspread.authorize(creds)

        # スプレッドシートをURLで開く
        spreadsheet = client.open_by_url(spreadsheet_url)
        worksheet = spreadsheet.get_worksheet(0) # 最初のシート (gid=0) を取得

        # 全てのデータを取得
        rows = worksheet.get_all_values()

        # ヘッダー行をスキップし、C列（投稿日、インデックス2）とD列（URL、インデックス3）を抽出
        news_data = [] # (URL, 投稿日時datetimeオブジェクト, A列の値, B列の値, C列の値) のリスト
        if len(rows) > 1: # ヘッダー行をスキップ
            for row_idx, row in enumerate(rows[1:], 2): # row_idxは1-basedで実際のシートの行番号
                # C列が投稿日、D列がURL
                if len(row) > 3 and row[3].startswith("http"): # URLが4列目（D列）にあり、httpで始まることを確認
                    try:
                        post_date_str = row[2] # ★C列の値 (投稿日、インデックス2)
                        post_datetime = None
                        if post_date_str:
                            # 投稿日の書式: YYYY/MM/DD HH:MM:SS
                            try:
                                post_datetime = datetime.strptime(post_date_str, "%Y/%m/%d %H:%M:%S")
                            except ValueError:
                                # もし秒までない場合も考慮
                                try:
                                    post_datetime = datetime.strptime(post_date_str, "%Y/%m/%d %H:%M")
                                except ValueError:
                                    print(f"警告: スプレッドシートの投稿日時形式を認識できませんでした (行 {row_idx}, 値: '{post_date_str}')。この行はスキップされます。")
                                    continue # パース失敗したらスキップ

                        news_data.append({
                            'url': row[3], # ★D列の値 (URL、インデックス3)
                            'post_datetime': post_datetime,
                            'original_A_col': row[0] if len(row) > 0 else '',
                            'original_B_col': row[1] if len(row) > 1 else '',
                            'original_C_col': row[2] if len(row) > 2 else '', # ★C列の値も保持
                            'original_row_index': row_idx # 後でコメント数を書き込むための元の行番号
                        })
                    except Exception as e:
                        print(f"スプレッドシートの行 ({row_idx}) 処理中にエラーが発生しました: {e} - 行データ: {row}")
                        continue # エラーが発生しても次の行へ進む
        return rows, news_data # 元の全行データと、抽出したニュースデータ

    except Exception as e:
        print(f"Google SheetsからのURL読み込みに失敗しました: {e}")
        return [], []

def upload_file_to_drive(file_path, drive_folder_id=None):
    """ファイルをGoogle Driveにアップロードする"""
    try:
        # PyDrive認証
        gauth = GoogleAuth()
        # サービスアカウントキーファイルを使用
        gauth.LoadCredentialsFile(SERVICE_ACCOUNT_FILE)
        if gauth.credentials is None:
            gauth.ServiceAuth() # サービスアカウント認証
        elif gauth.access_token_expired:
            gauth.Refresh()
        else:
            gauth.Authorize()

        drive = GoogleDrive(gauth)

        # ファイルメタデータの設定
        file_name = os.path.basename(file_path)
        file_metadata = {'title': file_name}
        if drive_folder_id:
            # フォルダIDが指定されている場合、親フォルダとして設定
            file_metadata['parents'] = [{"kind": "drive#fileLink", "id": drive_folder_id}]

        # ファイルを作成し、内容をアップロード
        f = drive.CreateFile(file_metadata)
        f.SetContentFile(file_path)
        f.Upload()

        print(f"✅ '{file_name}' を Google Drive にアップロードしました。ファイルID: {f['id']}")
        return f['id']
    except Exception as e:
        print(f"Google Driveへのアップロードに失敗しました: {e}")
        print(f"エラー詳細: {e}")
        return None

def main():
    # Google Sheetsからデータ読み込み
    all_sheet_rows, news_list_from_sheet = get_urls_from_google_sheet(SPREADSHEET_URL)
    if not news_list_from_sheet:
        print("❌ Google Sheetsから処理対象のURLを1件も読み込めませんでした。プログラムを終了します。")
        return

    # 日付フィルタリングロジック: 前日15:00から当日14:59の範囲
    # 現在時刻はJST 11:25:55 なので、今日の日付は 2025/07/30
    now_jst = datetime.now() # GitHub ActionsはUTCだが、ここでは実行時のタイムゾーンをJSTと仮定して処理
    # 実行日の前日15:00 (JST)
    start_time = (now_jst - timedelta(days=1)).replace(hour=15, minute=0, second=0, microsecond=0)
    # 実行日の当日14:59 (JST)
    end_time = now_jst.replace(hour=14, minute=59, second=59, microsecond=999999)

    # フィルタリングされたニュースデータ
    filtered_news_data = []
    print(f"フィルタリング期間: {start_time.strftime('%Y/%m/%d %H:%M')} ～ {end_time.strftime('%Y/%m/%d %H:%M')}")
    for news_item in news_list_from_sheet:
        post_datetime = news_item['post_datetime']
        if post_datetime and start_time <= post_datetime <= end_time:
            filtered_news_data.append(news_item)

    if not filtered_news_data:
        print(f"フィルタリング条件に合致するニュースはありませんでした。")
        return

    print(f"✅ フィルタリングされたニュース数: {len(filtered_news_data)} 件を処理します。")

    # Chrome WebDriverの設定
    options = Options()
    options.add_argument("--lang=ja-JP")
    options.add_argument("--start-maximized")
    # GitHub Actionsで実行する場合はヘッドレスモードを有効にする
    options.add_argument("--headless")
    options.add_argument("--no-sandbox") # Docker/CI環境で必要になる場合がある
    options.add_argument("--disable-dev-shm-usage") # Docker/CI環境で必要になる場合がある
    driver = webdriver.Chrome(options=options)

    # Excelワークブックの作成
    wb = Workbook()
    ws_input = wb.active
    ws_input.title = "input"

    # 元のスプレッドシートの内容を 'input' シートに書き込む
    for r_idx, row_data in enumerate(all_sheet_rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            ws_input.cell(row=r_idx, column=c_idx, value=val)

    # 'input' シートのヘッダーに「コメント件数」列を追加 (F列)
    # スプレッドシートの列数に応じて調整が必要（A=1, B=2, C=3, D=4, E=5, F=6）
    comment_count_col = max(len(all_sheet_rows[0]) if all_sheet_rows else 0, 5) + 1 # 既存列+1 (最低F列)
    ws_input.cell(row=1, column=comment_count_col, value="コメント件数")


    # フィルタリングされたニュースを順次処理
    for idx, news_item in enumerate(filtered_news_data, 1):
        url = news_item['url']
        original_row_index = news_item['original_row_index'] # 元のスプレッドシートでの行番号

        print(f"\n▶ ({idx}/{len(filtered_news_data)}) 処理中: {url}")

        # Excelシート名の生成
        # 最大31文字制限、特殊文字は避ける
        sheet_title_base = f"News_{idx}"
        if news_item['original_A_col']:
             # A列の値をシート名に含める場合、文字数と特殊文字に注意
             clean_a_col = "".join(c for c in news_item['original_A_col'] if c.isalnum() or c in ('_', '-'))
             sheet_title_base = f"{clean_a_col[:20]}_{idx}" # A列の先頭20文字とインデックス
        sheet_title = sheet_title_base[:31] # 最終的なシート名は31文字以内

        # シート名が重複しないように調整
        counter = 0
        final_sheet_title = sheet_title
        while final_sheet_title in wb.sheetnames:
            counter += 1
            final_sheet_title = f"{sheet_title[:28]}_{counter}" # 28文字 + '_XX'

        ws = wb.create_sheet(title=final_sheet_title)

        # ニュース記事の取得
        try:
            title, base_url, pages = get_news_pages(url, driver)
            ws.cell(row=1, column=1, value="タイトル")
            ws.cell(row=1, column=2, value=title)
            ws.cell(row=2, column=1, value="URL")
            ws.cell(row=2, column=2, value=base_url)

            # 記事本文をシートに書き込み (最大15ページ)
            for i, page_text in enumerate(pages[:15], 1):
                ws.cell(row=i + 2, column=1, value=page_text)

            # 記事セクションの空白行埋め
            for i in range(len(pages)+3, 18):
                ws.cell(row=i, column=1, value="")

        except Exception as e:
            ws.cell(row=1, column=1, value="エラー")
            ws.cell(row=2, column=1, value=f"記事取得失敗: {str(e)}")
            print(f"記事取得エラー ({url}): {e}")

        # コメントの取得
        try:
            comments = get_comments_pages(url, driver)
            start_row = 20
            ws.cell(row=start_row - 1, column=1, value="コメント本文")
            ws.cell(row=start_row - 1, column=2, value="投稿日時")
            ws.cell(row=start_row - 1, column=3, value="ユーザー名")

            for i, (text, dt, user) in enumerate(comments, start=start_row):
                ws.cell(row=i, column=1, value=text)
                ws.cell(row=i, column=2, value=dt)
                ws.cell(row=i, column=3, value=user)

            # コメント件数を計算して 'input' シートに書き込む
            comment_count = len(comments) if comments[0][0] != "コメントなし" else 0
            ws_input.cell(row=original_row_index, column=comment_count_col, value=comment_count)

        except Exception as e:
            ws.cell(row=20, column=1, value="コメント取得失敗")
            ws.cell(row=20, column=2, value=str(e))
            print(f"コメント取得エラー ({url}): {e}")
            ws_input.cell(row=original_row_index, column=comment_count_col, value="取得失敗")

    driver.quit() # ブラウザを閉じる
    wb.save(OUTPUT_FILE) # ローカルに一時保存

    print(f"\n✅ ローカルに一時保存完了！ファイル: {os.path.abspath(OUTPUT_FILE)}")

    # Google Driveにアップロード
    upload_file_to_drive(OUTPUT_FILE, DRIVE_FOLDER_ID)

    # ローカルの一時ファイルを削除 (オプション)
    try:
        os.remove(OUTPUT_FILE)
        print(f"ローカル一時ファイル '{OUTPUT_FILE}' を削除しました。")
    except OSError as e:
        print(f"ローカル一時ファイルの削除に失敗しました: {e}")

if __name__ == "__main__":
    main()
